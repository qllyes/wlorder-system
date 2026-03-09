"""
waybill_generator.py — 托运单 Excel 生成工具库

主要职责：
  1. 解析客户订单 Excel（parse_order_excel）—— 提取全量原始行数据
  2. 计算总重量（calc_total_weight）
  3. 金额转中文大写（num_to_chinese）
  4. 以「托运单.xlsx」为模板生成填充后的字节流（generate_waybill_excel）
"""
from __future__ import annotations

import io
import re
import datetime
from pathlib import Path


# ─── 固定配置（硬编码）────────────────────────────────────────────
WAYBILL_TEMPLATE = Path(__file__).parent / "托运单.xlsx"

FIXED_CONFIG: dict = {
    "发站":   "七箭啤酒厂",
    "发货方": "物流部",
    "发货电话": "13487819747",
    "查货电话": "13487819747",
    "业务电话": "李13016141852 /王15080730688",
}

# 规格中关键字 → 单重(kg) 默认映射（可被 db 中的配置覆盖）
DEFAULT_SPEC_WEIGHTS: dict[str, float] = {
    "305ml*12":  6.54,
    "305ml×12":  6.54,
    "650ml*6":   4.72,
    "650ml×6":   4.72,
    "650ml*12":  9.25,
    "650ml×12":  9.25,
    "1l*6":      7.14,
    "1L*6":      7.14,
    "750ml*6":   10.0,
    "750ml×6":   10.0,
    "20l/桶":    21.0,
    "20L/桶":    21.0,
}


def normalize_spec_text(text: str) -> str:
    t = (text or '').strip().lower()
    t = t.replace('×', '*').replace('x', '*').replace('＊', '*')
    t = re.sub(r'\s+', '', t)
    return t


def parse_spec_from_product_name(product_name: str) -> str:
    """从商品名称中提取规格片段，如 305ml*12 / 20L/桶。"""
    name = normalize_spec_text(product_name)
    patterns = [
        r'(\d+(?:\.\d+)?(?:ml|l)\*\d+)',
        r'(\d+(?:\.\d+)?(?:ml|l)/桶)',
        r'(\d+(?:\.\d+)?(?:ml|l)/箱)',
    ]
    for pat in patterns:
        m = re.search(pat, name, flags=re.IGNORECASE)
        if m:
            return m.group(1)
    return ''


def match_spec_weight(product_name: str, spec: str, spec_weights: dict[str, float]) -> tuple[str, float, str]:
    """返回(解析规格, 单重kg, 来源)。"""
    normalized_sw = {normalize_spec_text(k): float(v) for k, v in spec_weights.items()}
    parsed_spec = parse_spec_from_product_name(product_name) or normalize_spec_text(spec)
    if parsed_spec and parsed_spec in normalized_sw:
        return parsed_spec, normalized_sw[parsed_spec], 'spec_match'

    combined = normalize_spec_text(f"{product_name}{spec}")
    for key, w in normalized_sw.items():
        if key and key in combined:
            return key, w, 'spec_match'
    return parsed_spec, 0.0, 'unmatched'

# ─── 1. 解析订单 Excel ────────────────────────────────────────────

def parse_order_excel(file_path: str | Path) -> dict:
    """
    解析客户订单 Excel，提取完整原始行数据。

    返回结构：
    {
        "order_no": "20260302-12",
        "receiver_name": "仲其林",
        "receiver_phone": "13951240761",
        "receiver_address": "江苏省...193号",
        "products": [
            {
                "name": "百香果305ml*12",
                "spec": "箱",
                "qty": 80,
                "_raw": {"序号": 1, "发货日期": ..., "商品名称": "百香果305ml*12", ...}
            },
            ...
        ]
    }
    """
    import openpyxl

    file_path = Path(file_path)

    # 从文件名中提取托运单编号（数字+连字符部分，如 20260302-12）
    order_no = ""
    m = re.search(r"(\d{8}-\d+)", file_path.stem)
    if m:
        order_no = m.group(1)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # 第1行为表头，从第2行开始读取数据
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]

    # 建立列名→列索引映射（兼容中文列名）
    idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h:
            idx[str(h).strip()] = i

    # 用于收货人信息（取第一行数据行）
    receiver_name = ""
    receiver_phone = ""
    receiver_address = ""
    products: list[dict] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue

        # ── 构建当前行的原始完整字典（表头 → 值）──
        raw_row: dict = {}
        for col_idx, header in enumerate(headers):
            if header and col_idx < len(row):
                val = row[col_idx]
                # datetime 对象转为字符串以便 JSON 序列化
                if isinstance(val, (datetime.datetime, datetime.date)):
                    val = val.strftime("%Y-%m-%d")
                raw_row[str(header).strip()] = val if val is not None else ""

        def _g(key: str) -> str:
            i = idx.get(key, -1)
            v = row[i] if (i >= 0 and i < len(row)) else None
            return str(v).strip() if v is not None else ""

        if not receiver_name:
            receiver_name = _g("收货人")
            receiver_phone = _g("收货电话")
            receiver_address = _g("收货地址")

        # 兼容多种表头命名：优先新名称，fallback 旧名称
        name = _g("商品名称") or _g("品名")
        spec = _g("包装/规格") or _g("规格")
        qty_raw = _g("件数") or _g("数量")
        try:
            qty = int(float(qty_raw)) if qty_raw else 0
        except ValueError:
            qty = 0

        if name:
            parsed_spec = parse_spec_from_product_name(name)
            products.append({
                "name": name,
                "spec": spec,
                "qty": qty,
                "parsed_spec": parsed_spec,
                "_raw": raw_row,  # 嵌套存储，避免字段名碰撞
            })

    return {
        "order_no": order_no,
        "receiver_name": receiver_name,
        "receiver_phone": receiver_phone,
        "receiver_address": receiver_address,
        "products": products,
    }


# ─── 2. 计算总重量 ────────────────────────────────────────────────

def _match_spec_weight(product_name: str, spec: str, spec_weights: dict[str, float]) -> float:
    """尝试从品名或规格字段匹配单重（kg），找不到返回0.0。"""
    combined = (product_name + spec).lower()
    for key, w in spec_weights.items():
        if key.lower() in combined:
            return w
    return 0.0


def enrich_products_with_weight(
    products: list[dict],
    spec_weights: dict[str, float] | None = None,
) -> list[dict]:
    sw = {**DEFAULT_SPEC_WEIGHTS, **(spec_weights or {})}
    enriched: list[dict] = []
    for p in products:
        qty = int(p.get('qty', p.get('quantity', 0)) or 0)
        parsed_spec, unit_w, source = match_spec_weight(p.get('name', p.get('product_name', '')), p.get('spec', ''), sw)
        unit_weight = float(p.get('unit_weight_kg', unit_w) or 0)
        line_weight = float(p.get('line_weight_kg', round(unit_weight * qty, 3)) or 0)
        merged = {**p}
        merged.update({
            'parsed_spec': p.get('parsed_spec') or parsed_spec,
            'unit_weight_kg': unit_weight,
            'line_weight_kg': line_weight,
            'weight_source': p.get('weight_source') or source,
            'weight_locked': int(p.get('weight_locked', 0) or 0),
        })
        enriched.append(merged)
    return enriched


def calc_total_weight(
    products: list[dict],
    spec_weights: dict[str, float] | None = None,
) -> tuple[float, float]:
    """
    计算总件数和总重量（吨）。

    Returns:
        (total_qty, total_weight_t) 总件数、总吨数（1位小数）
    """
    sw = {**DEFAULT_SPEC_WEIGHTS, **(spec_weights or {})}
    total_qty = sum(p["qty"] for p in products)
    total_kg = 0.0
    for p in products:
        if p.get('line_weight_kg') is not None:
            total_kg += float(p.get('line_weight_kg') or 0)
        else:
            total_kg += p["qty"] * _match_spec_weight(p["name"], p.get("spec", ""), sw)
    total_t = round(total_kg / 1000, 1)
    return total_qty, total_t


# ─── 3. 金额转中文大写 ────────────────────────────────────────────

def num_to_chinese(amount: float) -> str:
    """将金额数字转换为中文大写，如 1050 → '壹仟零伍拾元整'。"""
    if amount <= 0:
        return "零元整"

    units_upper = ["", "拾", "佰", "仟", "万", "拾", "佰", "仟", "亿"]
    digits = ["零", "壹", "贰", "叁", "肆", "伍", "陆", "柒", "捌", "玖"]

    # 处理整数部分（精度取到元）
    int_part = int(round(amount))
    if int_part == 0:
        return "零元整"

    result = ""
    s = str(int_part)
    length = len(s)
    prev_zero = False

    for i, ch in enumerate(s):
        d = int(ch)
        unit = units_upper[length - 1 - i]
        if d == 0:
            prev_zero = True
        else:
            if prev_zero and result:
                result += "零"
            result += digits[d] + unit
            prev_zero = False

    result += "元整"
    # 万字位修正
    result = result.replace("亿万", "亿")
    return result


# ─── 4. 生成托运单 Excel ──────────────────────────────────────────

# 商品区行号：左侧6行（A-C列），右侧6行（D-F列）
_PRODUCT_LEFT_ROWS  = [7, 8, 9, 10, 11, 12]
_PRODUCT_RIGHT_ROWS = [7, 8, 9, 10, 11, 12]


def _extract_dao_zhan(address: str) -> str:
    """从收货地址提取到站信息（取省市级别）。"""
    m = re.match(r"(.{2,4}省|.{2,4}市|.{2,4}自治区)(.{2,5}市)?", address)
    if m:
        parts = [p for p in m.groups() if p]
        return "".join(parts)
    return address[:8]  # 兜底：取前8字


def generate_waybill_excel(
    order_data: dict,
    freight: float = 0.0,
    pickup_method: str = "送货上门",
    payment_method: str = "现付",
    spec_weights: dict[str, float] | None = None,
) -> bytes:
    """
    以「托运单.xlsx」为模板，填充数据后返回字节流。
    策略：先取消全部合并 → 写入数据 → 恢复合并，避免 MergedCell read-only 报错。
    """
    import openpyxl

    wb = openpyxl.load_workbook(WAYBILL_TEMPLATE)
    ws = wb.active

    # ── Step 1：记录并临时取消所有合并单元格 ──
    merged_ranges = [str(r) for r in list(ws.merged_cells.ranges)]
    for mr in merged_ranges:
        ws.unmerge_cells(mr)

    # ── Step 2：计算数据 ──
    products = order_data.get("products", [])[:12]
    total_qty, total_weight_t = calc_total_weight(products, spec_weights)
    amount_cn = num_to_chinese(freight)
    dao_zhan = _extract_dao_zhan(order_data.get("receiver_address", ""))

    # ── Step 3：填写各字段（写入合并区左上角单元格） ──
    today = datetime.date.today()
    ws["A2"] = f"托运日期:  {today.year}年{today.month}月{today.day}日"
    ws["F2"] = f"X{order_data.get('order_no', '')}"   # 合并 F2:H2 → 写 F2

    ws["B3"] = FIXED_CONFIG["发站"]
    ws["D3"] = FIXED_CONFIG["发货方"]
    ws["F3"] = FIXED_CONFIG["发货电话"]               # 合并 F3:H3 → 写 F3

    ws["B4"] = dao_zhan
    ws["D4"] = order_data.get("receiver_name", "")
    ws["F4"] = order_data.get("receiver_phone", "")   # 合并 F4:H4 → 写 F4
    ws["B5"] = order_data.get("receiver_address", "")

    # 商品区（左侧）
    for i, p in enumerate(products[:6]):
        row = _PRODUCT_LEFT_ROWS[i]
        ws[f"A{row}"] = p["name"]
        ws[f"B{row}"] = p.get("spec", "")
        ws[f"C{row}"] = p["qty"] if p["qty"] else ""

    # 商品区（右侧，第7-12个）
    for i, p in enumerate(products[6:12]):
        row = _PRODUCT_RIGHT_ROWS[i]
        ws[f"D{row}"] = p["name"]
        ws[f"E{row}"] = p.get("spec", "")
        ws[f"F{row}"] = p["qty"] if p["qty"] else ""

    # 汇总区
    ws["E11"] = total_qty                              # 合并 E11:F11 → 写 E11
    ws["E12"] = f"{total_weight_t} 吨"                # 合并 E12:F12 → 写 E12
    ws["G6"]  = freight if freight else ""             # 合并 G6:H6  → 运费 写 G6
    ws["A14"] = amount_cn                              # 合并 A14:H14 → 大写 写 A14

    # 取货/付款方式
    v_pickup  = "√" if pickup_method  == "送货上门" else " "
    v_ziti    = "√" if pickup_method  == "自提"    else " "
    v_tif     = "√" if payment_method == "提付"    else " "
    v_xian    = "√" if payment_method == "现付"    else " "
    ws["A16"] = (
        f"取货方式：送货上门 ({v_pickup})  自提 ({v_ziti})"
        f"    付款方式：提付 ({v_tif})  现付 ({v_xian})"
        f"       查货电话：{FIXED_CONFIG['查货电话']}  业务电话：{FIXED_CONFIG['业务电话']}"
    )

    # ── Step 4：恢复合并单元格 ──
    for mr in merged_ranges:
        ws.merge_cells(mr)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
    import openpyxl
